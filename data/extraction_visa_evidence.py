from pathlib import Path
from openpyxl import load_workbook
import csv


# The Excel files are stored in the same folder as this Python file
folder = Path(__file__).parent


# Original dataset filenames
visitor_file = folder / (
    "bp0018l-visitor-visas-granted-quarterly-report-"
    "as-at-2026-06-30-v100 (1).xlsx"
)

student_lodged_file = folder / (
    "bp0015l-student-visas-lodged-report-locked-"
    "at-2026-06-30-v100.xlsx"
)

student_granted_file = folder / (
    "bp0015l-student-visas-granted-report-locked-"
    "at-2026-06-30-v100.xlsx"
)

temporary_holders_file = folder / (
    "bp0019l-number-of-temporary-visa-holders-in-"
    "australia-at-2026-06-30-v100.xlsx"
)


# Check that all four files can be found
files = [
    visitor_file,
    student_lodged_file,
    student_granted_file,
    temporary_holders_file
]

for file in files:
    if not file.exists():
        print("Could not find:", file.name)
        quit()

print("All four visa datasets were found.")
print()


# A simple function for displaying each piece of evidence
def show_evidence(passage_id, question, source, location, evidence):
    print("=" * 70)
    print("Passage ID:", passage_id)
    print("Question:", question)
    print("Source:", source)
    print("Sheet and cells:", location)
    print("Evidence:", evidence)
    print()



# DATASET 1: Visitor visas granted


visitor_workbook = load_workbook(
    visitor_file,
    read_only=True,
    data_only=True
)

granted_sheet = visitor_workbook["Granted"]
notes_sheet = visitor_workbook["Explanatory Notes"]
terminology_sheet = visitor_workbook["Data Items and Terminology Used"]


# Extract the ETA description directly from cells B16 to B18
eta_parts = []

for row_number in range(16, 19):
    value = notes_sheet[f"B{row_number}"].value

    if value is not None:
        eta_parts.append(str(value).strip())

eta_evidence = " ".join(eta_parts)

show_evidence(
    "VP001",
    "How long can an Electronic Travel Authority holder stay during each visit?",
    visitor_file.name,
    "Explanatory Notes!B16:B18",
    eta_evidence
)


# Extract the Frequent Traveller stream description from C12 to C14
frequent_traveller_parts = []

for row_number in range(12, 15):
    value = terminology_sheet[f"C{row_number}"].value

    if value is not None:
        frequent_traveller_parts.append(str(value).strip())

frequent_traveller_evidence = " ".join(frequent_traveller_parts)

show_evidence(
    "VP002",
    "What are the stay conditions for the Frequent Traveller stream?",
    visitor_file.name,
    "Data Items and Terminology Used!C12:C14",
    frequent_traveller_evidence
)


# Extract visitor visa grant totals
visitor_grand_total = granted_sheet["W23"].value
tourist_total = granted_sheet["W18"].value
business_total = granted_sheet["W22"].value
sponsored_family_total = granted_sheet["W14"].value

show_evidence(
    "VP003",
    "How many visitor visas were granted in total in 2025-26 to 30 June 2026?",
    visitor_file.name,
    "Granted!W23",
    f"Grand Total | 2025-26 to 30 June 2026 | {visitor_grand_total:,}"
)

show_evidence(
    "VP004",
    "How many Tourist stream visitor visas were granted in 2025-26 to 30 June 2026?",
    visitor_file.name,
    "Granted!W18",
    f"Tourist Total | 2025-26 to 30 June 2026 | {tourist_total:,}"
)

show_evidence(
    "VP005",
    "How many Business stream visitor visas were granted in 2025-26 to 30 June 2026?",
    visitor_file.name,
    "Granted!W22",
    f"Business Total | 2025-26 to 30 June 2026 | {business_total:,}"
)

show_evidence(
    "VP006",
    "How many Sponsored Family stream visitor visas were granted in 2025-26 to 30 June 2026?",
    visitor_file.name,
    "Granted!W14",
    f"Sponsored Family | 2025-26 to 30 June 2026 | {sponsored_family_total:,}"
)

visitor_workbook.close()



# DATASET 2: Student visa applications lodged


lodged_workbook = load_workbook(
    student_lodged_file,
    read_only=True,
    data_only=True
)

# This sheet name contains a space at the end
lodged_sheet = lodged_workbook["Lodged (Month) "]

student_lodged_total = lodged_sheet["W31"].value
higher_education_lodged = lodged_sheet["W16"].value

show_evidence(
    "VP007",
    "How many student visa applications were lodged in total in 2025-26 to 30 June 2026?",
    student_lodged_file.name,
    "Lodged (Month) !W31",
    f"Grand Total | 2025-26 to 30 June 2026 | {student_lodged_total:,}"
)

show_evidence(
    "VP008",
    "How many primary Higher Education student visa applications were lodged in 2025-26 to 30 June 2026?",
    student_lodged_file.name,
    "Lodged (Month) !W16",
    f"Primary Higher Education | 2025-26 to 30 June 2026 | {higher_education_lodged:,}"
)

lodged_workbook.close()



# DATASET 3: Student visas granted

student_granted_workbook = load_workbook(
    student_granted_file,
    read_only=True,
    data_only=True
)

student_granted_sheet = student_granted_workbook["Granted (Month)"]

student_granted_total = student_granted_sheet["W32"].value
higher_education_granted = student_granted_sheet["W17"].value

show_evidence(
    "VP009",
    "How many student visas were granted in total in 2025-26 to 30 June 2026?",
    student_granted_file.name,
    "Granted (Month)!W32",
    f"Grand Total | 2025-26 to 30 June 2026 | {student_granted_total:,}"
)

show_evidence(
    "VP010",
    "How many primary Higher Education student visas were granted in 2025-26 to 30 June 2026?",
    student_granted_file.name,
    "Granted (Month)!W17",
    f"Primary Higher Education | 2025-26 to 30 June 2026 | {higher_education_granted:,}"
)

student_granted_workbook.close()



# DATASET 4: Temporary visa holders in Australia


holders_workbook = load_workbook(
    temporary_holders_file,
    read_only=True,
    data_only=True
)

holders_sheet = holders_workbook["Visa Holders"]

visitor_holders = holders_sheet["W20"].value
student_holders = holders_sheet["W15"].value

show_evidence(
    "VP011",
    "How many Visitor visa holders were in Australia on 30 June 2026?",
    temporary_holders_file.name,
    "Visa Holders!W20",
    f"Visitor visa holders | 30 June 2026 | {visitor_holders:,}"
)

show_evidence(
    "VP012",
    "How many Student visa holders were in Australia on 30 June 2026?",
    temporary_holders_file.name,
    "Visa Holders!W15",
    f"Student visa holders | 30 June 2026 | {student_holders:,}"
)

holders_workbook.close()


print("=" * 70)
print("Finished extracting 12 pieces of visa evidence.")


# Store the 12 extracted passages in Walert's collection format
passages = [
    ["VP001", eta_evidence],
    ["VP002", frequent_traveller_evidence],
    [
        "VP003",
        f"Grand Total | 2025-26 to 30 June 2026 | {visitor_grand_total:,}"
    ],
    [
        "VP004",
        f"Tourist Total | 2025-26 to 30 June 2026 | {tourist_total:,}"
    ],
    [
        "VP005",
        f"Business Total | 2025-26 to 30 June 2026 | {business_total:,}"
    ],
    [
        "VP006",
        f"Sponsored Family | 2025-26 to 30 June 2026 | "
        f"{sponsored_family_total:,}"
    ],
    [
        "VP007",
        f"Grand Total | 2025-26 to 30 June 2026 | "
        f"{student_lodged_total:,}"
    ],
    [
        "VP008",
        f"Primary Higher Education | 2025-26 to 30 June 2026 | "
        f"{higher_education_lodged:,}"
    ],
    [
        "VP009",
        f"Grand Total | 2025-26 to 30 June 2026 | "
        f"{student_granted_total:,}"
    ],
    [
        "VP010",
        f"Primary Higher Education | 2025-26 to 30 June 2026 | "
        f"{higher_education_granted:,}"
    ],
    [
        "VP011",
        f"Visitor visa holders | 30 June 2026 | {visitor_holders:,}"
    ],
    [
        "VP012",
        f"Student visa holders | 30 June 2026 | {student_holders:,}"
    ]
]


# Create collection.csv in the same folder
collection_file = folder / "collection.csv"

with open(collection_file, "w", newline="", encoding="utf-8") as file:
    writer = csv.writer(file)

    #required column names
    writer.writerow(["passage_id", "passage"])

    # Add all 12 passages
    writer.writerows(passages)

print("Created:", collection_file.name)


# Store the test questions in Walert's topics format
topics = [
    [
        "VS01",
        "Electronic Travel Authority stay length",
        "VS01Q01",
        "How long can an Electronic Travel Authority holder stay during each visit?"
    ],
    [
        "VS02",
        "Frequent Traveller stream stay conditions",
        "VS02Q01",
        "What are the stay conditions for the Frequent Traveller stream?"
    ],
    [
        "VS03",
        "Total visitor visas granted",
        "VS03Q01",
        "How many visitor visas were granted in total in 2025-26 to 30 June 2026?"
    ],
    [
        "VS04",
        "Tourist stream visitor visas granted",
        "VS04Q01",
        "How many Tourist stream visitor visas were granted in 2025-26 to 30 June 2026?"
    ],
    [
        "VS05",
        "Business stream visitor visas granted",
        "VS05Q01",
        "How many Business stream visitor visas were granted in 2025-26 to 30 June 2026?"
    ],
    [
        "VS06",
        "Sponsored Family visitor visas granted",
        "VS06Q01",
        "How many Sponsored Family stream visitor visas were granted in 2025-26 to 30 June 2026?"
    ],
    [
        "VS07",
        "Total student visa applications lodged",
        "VS07Q01",
        "How many student visa applications were lodged in total in 2025-26 to 30 June 2026?"
    ],
    [
        "VS08",
        "Higher Education student visa applications lodged",
        "VS08Q01",
        "How many primary Higher Education student visa applications were lodged in 2025-26 to 30 June 2026?"
    ],
    [
        "VS09",
        "Total student visas granted",
        "VS09Q01",
        "How many student visas were granted in total in 2025-26 to 30 June 2026?"
    ],
    [
        "VS10",
        "Higher Education student visas granted",
        "VS10Q01",
        "How many primary Higher Education student visas were granted in 2025-26 to 30 June 2026?"
    ],
    [
        "VS11",
        "Visitor visa holders in Australia",
        "VS11Q01",
        "How many Visitor visa holders were in Australia on 30 June 2026?"
    ],
    [
        "VS12",
        "Student visa holders in Australia",
        "VS12Q01",
        "How many Student visa holders were in Australia on 30 June 2026?"
    ]
]


# Create topics.csv in the same folder
topics_file = folder / "topics.csv"

with open(topics_file, "w", newline="", encoding="utf-8") as file:
    writer = csv.writer(file)

    # These are the column names used by Walert
    writer.writerow(["topic_id", "Topic", "question_id", "question"])
    writer.writerows(topics)

print("Created:", topics_file.name)





# Create the ground-truth mappings
groundtruth = []

# The topics and passages are in the same order
for number in range(len(topics)):
    topic_id = topics[number][0]
    topic_name = topics[number][1]

    passage_id = passages[number][0]
    passage_text = passages[number][1]

    groundtruth.append([
        topic_id,
        topic_name,
        passage_id,
        passage_text,
        2
    ])


# Create groundtruth.csv
groundtruth_file = folder / "groundtruth.csv"

with open(groundtruth_file, "w", newline="", encoding="utf-8") as file:
    writer = csv.writer(file)

    writer.writerow([
        "topic_id",
        "topic",
        "passage_id",
        "passage",
        "relevance_judgment"
    ])

    writer.writerows(groundtruth)

print("Created:", groundtruth_file.name)

# Create the expected answers for each test question
gold_summaries = [
    [
        "VS01Q01",
        "VP001",
        eta_evidence,
        "VP001"
    ],
    [
        "VS02Q01",
        "VP002",
        frequent_traveller_evidence,
        "VP002"
    ],
    [
        "VS03Q01",
        "VP003",
        f"A total of {visitor_grand_total:,} visitor visas were granted "
        f"in 2025-26 to 30 June 2026.",
        "VP003"
    ],
    [
        "VS04Q01",
        "VP004",
        f"A total of {tourist_total:,} Tourist stream visitor visas "
        f"were granted in 2025-26 to 30 June 2026.",
        "VP004"
    ],
    [
        "VS05Q01",
        "VP005",
        f"A total of {business_total:,} Business stream visitor visas "
        f"were granted in 2025-26 to 30 June 2026.",
        "VP005"
    ],
    [
        "VS06Q01",
        "VP006",
        f"A total of {sponsored_family_total:,} Sponsored Family stream "
        f"visitor visas were granted in 2025-26 to 30 June 2026.",
        "VP006"
    ],
    [
        "VS07Q01",
        "VP007",
        f"A total of {student_lodged_total:,} student visa applications "
        f"were lodged in 2025-26 to 30 June 2026.",
        "VP007"
    ],
    [
        "VS08Q01",
        "VP008",
        f"A total of {higher_education_lodged:,} primary Higher Education "
        f"student visa applications were lodged in 2025-26 to "
        f"30 June 2026.",
        "VP008"
    ],
    [
        "VS09Q01",
        "VP009",
        f"A total of {student_granted_total:,} student visas were granted "
        f"in 2025-26 to 30 June 2026.",
        "VP009"
    ],
    [
        "VS10Q01",
        "VP010",
        f"A total of {higher_education_granted:,} primary Higher Education "
        f"student visas were granted in 2025-26 to 30 June 2026.",
        "VP010"
    ],
    [
        "VS11Q01",
        "VP011",
        f"There were {visitor_holders:,} Visitor visa holders in Australia "
        f"on 30 June 2026.",
        "VP011"
    ],
    [
        "VS12Q01",
        "VP012",
        f"There were {student_holders:,} Student visa holders in Australia "
        f"on 30 June 2026.",
        "VP012"
    ]
]


# Create gold_summaries.csv
gold_summaries_file = folder / "gold_summaries.csv"

with open(
    gold_summaries_file,
    "w",
    newline="",
    encoding="utf-8"
) as file:
    writer = csv.writer(file)

    writer.writerow([
        "question_id",
        "summary_id",
        "summary",
        "passage_id"
    ])

    writer.writerows(gold_summaries)

print("Created:", gold_summaries_file.name)





# Check that the Walert answer-key files are consistent
print()
print("Checking the answer-key files...")


# Collect all question and passage IDs
question_ids = []

for topic in topics:
    question_ids.append(topic[2])


passage_ids = []

for passage in passages:
    passage_ids.append(passage[0])


# Check that there are 12 questions and 12 passages
assert len(question_ids) == 12, "There should be 12 questions."
assert len(passage_ids) == 12, "There should be 12 passages."


# Check for duplicate IDs
assert len(question_ids) == len(set(question_ids)), (
    "There are duplicate question IDs."
)

assert len(passage_ids) == len(set(passage_ids)), (
    "There are duplicate passage IDs."
)


# Collect all topic IDs
topic_ids = []

for topic in topics:
    topic_ids.append(topic[0])


# Check every ground-truth mapping
for mapping in groundtruth:
    topic_id = mapping[0]
    passage_id = mapping[2]
    relevance = mapping[4]

    assert topic_id in topic_ids, (
        f"{topic_id} is missing from topics.csv"
    )

    assert passage_id in passage_ids, (
        f"{passage_id} is missing from collection.csv"
    )

    assert relevance in [1, 2], (
        f"{topic_id} has an invalid relevance score."
    )


# Check that the visa preparation files were created
answer_key_files = [
    collection_file,
    topics_file,
    groundtruth_file,
    gold_summaries_file
]

for answer_key_file in answer_key_files:
    assert answer_key_file.exists(), (
        f"{answer_key_file.name} was not created."
    )


print("Validation passed.")
print("12 questions were found.")
print("12 passages were found.")
print("All topic and passage IDs match.")
print("All visa preparation files exist.")
print("Run Walert's data.py next to create qrels.txt.")


# Record where each passage came from
evidence_sources = [
    [
        "VP001",
        visitor_file.name,
        "Explanatory Notes",
        "B16:B18"
    ],
    [
        "VP002",
        visitor_file.name,
        "Data Items and Terminology Used",
        "C12:C14"
    ],
    [
        "VP003",
        visitor_file.name,
        "Granted",
        "W23"
    ],
    [
        "VP004",
        visitor_file.name,
        "Granted",
        "W18"
    ],
    [
        "VP005",
        visitor_file.name,
        "Granted",
        "W22"
    ],
    [
        "VP006",
        visitor_file.name,
        "Granted",
        "W14"
    ],
    [
        "VP007",
        student_lodged_file.name,
        "Lodged (Month) ",
        "W31"
    ],
    [
        "VP008",
        student_lodged_file.name,
        "Lodged (Month) ",
        "W16"
    ],
    [
        "VP009",
        student_granted_file.name,
        "Granted (Month)",
        "W32"
    ],
    [
        "VP010",
        student_granted_file.name,
        "Granted (Month)",
        "W17"
    ],
    [
        "VP011",
        temporary_holders_file.name,
        "Visa Holders",
        "W20"
    ],
    [
        "VP012",
        temporary_holders_file.name,
        "Visa Holders",
        "W15"
    ]
]


# Create the evidence source record
evidence_sources_file = folder / "evidence_sources.csv"

with open(
    evidence_sources_file,
    "w",
    newline="",
    encoding="utf-8"
) as file:
    writer = csv.writer(file)

    writer.writerow([
        "passage_id",
        "source_file",
        "sheet_name",
        "cell_or_range"
    ])

    writer.writerows(evidence_sources)

print("Created:", evidence_sources_file.name)
